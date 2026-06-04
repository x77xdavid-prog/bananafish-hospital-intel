# -*- coding: utf-8 -*-
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl

RAW = os.path.join(os.path.dirname(__file__), "raw")
hosp = os.path.join(RAW, "1.병원정보서비스(2026.3.).xlsx")

wb = openpyxl.load_workbook(hosp, read_only=True, data_only=True)
print("SHEETS:", wb.sheetnames)
ws = wb[wb.sheetnames[0]]

rows = ws.iter_rows(values_only=True)
header = next(rows)
print("NCOLS:", len(header))
for i, h in enumerate(header):
    print(f"  [{i}] {h}")

print("---- 2 sample rows ----")
for n, r in enumerate(rows):
    print(n, list(r))
    if n >= 1:
        break
wb.close()
