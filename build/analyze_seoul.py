# -*- coding: utf-8 -*-
import sys, io, os
from collections import Counter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl

RAW = os.path.join(os.path.dirname(__file__), "raw")
hosp = os.path.join(RAW, "1.병원정보서비스(2026.3.).xlsx")
dept = os.path.join(RAW, "5.의료기관별상세정보서비스_03_진료과목정보(2026.3.).xlsx")

wb = openpyxl.load_workbook(hosp, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = ws.iter_rows(values_only=True)
header = next(rows)
# indices
I_SIDO, I_GU, I_CL, I_X, I_Y = 5, 7, 3, 28, 29
total_seoul = 0
coord_ok = 0
gu = Counter()
cl = Counter()
cl_seoul = Counter()
for r in rows:
    if r[I_SIDO] == '서울':
        total_seoul += 1
        gu[r[I_GU]] += 1
        cl_seoul[r[I_CL]] += 1
        if r[I_X] not in (None, '', 0) and r[I_Y] not in (None, '', 0):
            coord_ok += 1
wb.close()

print("=== SEOUL TOTALS ===")
print("total_seoul:", total_seoul)
print("coord_ok:", coord_ok, f"({coord_ok*100//max(total_seoul,1)}%)")
print("\n=== per GU (sorted) ===")
for k, v in sorted(gu.items(), key=lambda x: -x[1]):
    print(f"  {v:6d}  {k}")
print("\n=== per 종별 (Seoul) ===")
for k, v in sorted(cl_seoul.items(), key=lambda x: -x[1]):
    print(f"  {v:6d}  {k}")

# dept file structure
print("\n=== DEPT FILE ===")
wb2 = openpyxl.load_workbook(dept, read_only=True, data_only=True)
ws2 = wb2[wb2.sheetnames[0]]
r2 = ws2.iter_rows(values_only=True)
h2 = next(r2)
print("SHEETS:", wb2.sheetnames)
print("NCOLS:", len(h2))
for i, h in enumerate(h2):
    print(f"  [{i}] {h}")
print("---- 3 sample rows ----")
for n, rr in enumerate(r2):
    print(n, list(rr))
    if n >= 2:
        break
wb2.close()
