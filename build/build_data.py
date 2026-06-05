# -*- coding: utf-8 -*-
"""
바나나피쉬 병원 인텔리전스 — data.js 생성기
출처: 건강보험심사평가원 '전국 병의원 및 약국 현황' (공공누리 출처표시, 영리·가공 허용)
입력: raw/1.병원정보서비스(2026.3.).xlsx, raw/5...진료과목정보(2026.3.).xlsx
출력: ../assets/data.js
"""
import sys, io, os, json, hashlib, datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl

HERE = os.path.dirname(__file__)
RAW = os.path.join(HERE, "raw")
OUT = os.path.normpath(os.path.join(HERE, "..", "assets", "data.js"))
HOSP = os.path.join(RAW, "1.병원정보서비스(2026.3.).xlsx")
DEPT = os.path.join(RAW, "5.의료기관별상세정보서비스_03_진료과목정보(2026.3.).xlsx")

# ── 수록 범위 (여기만 바꾸면 확장) ───────────────────────────
TARGET_SIDO = None   # None = 전국 / "서울" 등으로 한정 가능
TARGET_GU = None   # None = 시도 내 전체 자치구
SRC_LABEL = "건강보험심사평가원 · 전국 병의원 및 약국 현황(2026.3.)"
DATA_VER = "2026.3"
# ────────────────────────────────────────────────────────────

# 컬럼 인덱스 (병원정보서비스)
C_YK, C_NAME, C_CLCD, C_CL = 0, 1, 2, 3
C_SIDO, C_SIDOCD, C_GU = 5, 4, 7
C_EMD, C_ZIP, C_ADDR, C_TEL, C_URL = 8, 9, 10, 11, 12
C_OPEN, C_DRTOT = 13, 14
C_MD_GEN, C_MD_INT, C_MD_RES, C_MD_SP = 15, 16, 17, 18      # 의과
C_DN_GEN, C_DN_INT, C_DN_RES, C_DN_SP = 19, 20, 21, 22      # 치과
C_HN_GEN, C_HN_INT, C_HN_RES, C_HN_SP = 23, 24, 25, 26      # 한방
C_X, C_Y = 28, 29


def sid(ykiho):
    return hashlib.md5(str(ykiho).encode("utf-8")).hexdigest()[:8]


def fmt_open(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if v in (None, ""):
        return None
    s = str(v)
    if len(s) == 8 and s.isdigit():
        return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"
    return s


def ival(v):
    try:
        return int(v) if v not in (None, "") else 0
    except (TypeError, ValueError):
        return 0


# ── Pass 1: 병원 기본정보 ────────────────────────────────────
print("[1/3] 병원 기본정보 읽는 중...")
wb = openpyxl.load_workbook(HOSP, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
it = ws.iter_rows(values_only=True)
next(it)  # header

records = {}     # id -> record dict
yk2id = {}       # ykiho -> id
for r in it:
    if TARGET_SIDO is not None and r[C_SIDO] != TARGET_SIDO:
        continue
    if TARGET_GU is not None and r[C_GU] not in TARGET_GU:
        continue
    x, y = r[C_X], r[C_Y]
    if x in (None, "", 0) or y in (None, "", 0):
        continue
    yk = r[C_YK]
    _id = sid(yk)
    yk2id[yk] = _id
    # 의사 구성 요약
    md_sp = ival(r[C_MD_SP]); dn_sp = ival(r[C_DN_SP]); hn_sp = ival(r[C_HN_SP])
    gen = ival(r[C_MD_GEN]) + ival(r[C_DN_GEN]) + ival(r[C_HN_GEN])
    intn = ival(r[C_MD_INT]) + ival(r[C_DN_INT]) + ival(r[C_HN_INT])
    res = ival(r[C_MD_RES]) + ival(r[C_DN_RES]) + ival(r[C_HN_RES])
    records[_id] = {
        "id": _id,
        "name": (r[C_NAME] or "").strip(),
        "cl": r[C_CL],
        "sido": r[C_SIDO],
        "gu": r[C_GU],
        "emd": r[C_EMD],
        "addr": (r[C_ADDR] or "").strip(),
        "tel": (r[C_TEL] or "").strip(),
        "url": (r[C_URL] or "").strip(),
        "open": fmt_open(r[C_OPEN]),
        "dr": ival(r[C_DRTOT]),
        "sp": md_sp + dn_sp + hn_sp,   # 전문의 합
        "gen": gen, "intn": intn, "res": res,
        "lat": round(float(y), 6),
        "lng": round(float(x), 6),
        "depts": [],
    }
wb.close()
print(f"      → 대상 {len(records):,}곳 (좌표 보유)")

# ── Pass 2: 진료과목 조인 ────────────────────────────────────
print("[2/3] 진료과목 조인 중...")
wb2 = openpyxl.load_workbook(DEPT, read_only=True, data_only=True)
ws2 = wb2[wb2.sheetnames[0]]
it2 = ws2.iter_rows(values_only=True)
next(it2)
DC_YK, DC_NM, DC_SP = 0, 3, 4
joined = 0
for r in it2:
    yk = r[DC_YK]
    _id = yk2id.get(yk)
    if not _id:
        continue
    nm = r[DC_NM]
    if not nm:
        continue
    records[_id]["depts"].append([nm, ival(r[DC_SP])])
    joined += 1
wb2.close()
print(f"      → 과목 매핑 {joined:,}건")

# ── Pass 3: data.js 출력 ────────────────────────────────────
print("[3/3] data.js 쓰는 중...")
recs = list(records.values())
recs.sort(key=lambda d: (d["gu"], d["name"]))

# 종별 분포
from collections import Counter
cl_cnt = Counter(d["cl"] for d in recs)
gu_cnt = Counter(d["gu"] for d in recs)

# 수록 범위 라벨 — 설정된 범위에서 자동 도출(하드코딩 금지)
if TARGET_GU:
    REGION = (TARGET_SIDO or "") + " " + " · ".join(sorted(TARGET_GU))
elif TARGET_SIDO:
    REGION = TARGET_SIDO + " 전체"
else:
    REGION = "전국 " + str(len({d["sido"] for d in recs})) + "개 시도"

META = {
    "source": SRC_LABEL,
    "license": "공공누리 출처표시(제1유형) · 영리·가공 허용",
    "version": DATA_VER,
    "region": REGION,
    "generated": datetime.date.today().isoformat(),  # 빌드 시각 자동 기록
    "total": len(recs),
    "byType": dict(cl_cnt.most_common()),
    "byGu": dict(gu_cnt.most_common()),
}

def jx(o):
    return json.dumps(o, ensure_ascii=False, separators=(",", ":"))

with io.open(OUT, "w", encoding="utf-8") as f:
    f.write("// 자동 생성 — build/build_data.py · 출처: " + SRC_LABEL + "\n")
    f.write("// 라이선스: 공공누리 출처표시(제1유형) · 영리/가공 허용\n")
    f.write("// 레코드 스키마는 app.js 가 객체 키로 직접 사용\n")
    f.write("window.HII_META = " + jx(META) + ";\n")
    f.write("window.HII_HOSPITALS = [\n")
    for d in recs:
        f.write(jx(d) + ",\n")
    f.write("];\n")
    f.write("window.HII_DEEP = window.HII_DEEP || {};\n")

print(f"      → {OUT}")
print(f"      → {len(recs):,} records, {os.path.getsize(OUT)/1024/1024:.2f} MB")

# 심층 후보(대형·정보풍부 기관) 출력
print("\n=== 심층 도시에 후보 (병원급 이상) ===")
big = [d for d in recs if d["cl"] in ("상급종합", "종합병원", "병원", "요양병원", "한방병원", "치과병원", "정신병원")]
big.sort(key=lambda d: -d["dr"])
for d in big[:30]:
    print(f"  {d['dr']:4d}의사 [{d['cl']}] {d['name']} ({d['gu']}) id={d['id']}")
